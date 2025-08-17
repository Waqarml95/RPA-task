"""Excel writer utilities for managing workbook operations."""

import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from utils.config_loader import settings
from utils.logging_utils import execution_logger


class ExcelWriter:
    """Excel writer for managing workbook operations."""

    def __init__(self, filepath: Optional[Union[str, Path]] = None):
        """Initialize Excel writer."""
        if filepath:
            self.filepath = Path(filepath)
        else:
            output_dir = Path(settings.output.base_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            self.filepath = output_dir / settings.output.excel.filename

        self.workbook: Optional[Workbook] = None
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        """Ensure workbook exists, create if not."""
        if self.filepath.exists():
            try:
                self.workbook = load_workbook(self.filepath)
                execution_logger.log_action("load_workbook", str(self.filepath))
            except Exception as e:
                execution_logger.logger.warning(
                    f"Could not load existing workbook: {e}. Creating new one."
                )
                self.workbook = Workbook()
                # Remove default sheet
                if "Sheet" in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook["Sheet"])
        else:
            self.workbook = Workbook()
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])

    def write_df(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        start_row: int = 1,
        include_index: bool = False,
        header_style: bool = True,
    ) -> None:
        """
        Write DataFrame to sheet, replacing existing content.

        Args:
            df: DataFrame to write
            sheet_name: Name of the sheet
            start_row: Starting row (1-based)
            include_index: Whether to include DataFrame index
            header_style: Whether to apply header styling
        """
        if not self.workbook:
            self._ensure_workbook()

        # Remove sheet if exists
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])

        # Create new sheet
        ws = self.workbook.create_sheet(sheet_name)

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=include_index, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply header style
                if header_style and r_idx == start_row:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Autosize columns if configured
        if settings.output.excel.autosize_columns:
            self._autosize_columns(ws)

        # Freeze panes if configured
        if settings.output.excel.freeze_panes:
            ws.freeze_panes = settings.output.excel.freeze_panes

        # Save workbook
        self.save()

        execution_logger.log_action(
            "write_dataframe",
            sheet_name,
            rows=len(df),
            columns=len(df.columns),
        )

    def append_df(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        include_header: bool = False,
        include_index: bool = False,
    ) -> None:
        """
        Append DataFrame to existing sheet.

        Args:
            df: DataFrame to append
            sheet_name: Name of the sheet
            include_header: Whether to include column headers
            include_index: Whether to include DataFrame index
        """
        if not self.workbook:
            self._ensure_workbook()

        # Get or create sheet
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            start_row = ws.max_row + 1
        else:
            ws = self.workbook.create_sheet(sheet_name)
            start_row = 1
            include_header = True  # Force header for new sheet

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=include_index, header=include_header), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Autosize columns
        if settings.output.excel.autosize_columns:
            self._autosize_columns(ws)

        # Save workbook
        self.save()

        execution_logger.log_action(
            "append_dataframe",
            sheet_name,
            rows=len(df),
            start_row=start_row,
        )

    def embed_image(
        self,
        image_path: Union[str, Path],
        sheet_name: str,
        cell: str = "A1",
        width: Optional[int] = None,
        height: Optional[int] = None,
    ) -> None:
        """
        Embed image in worksheet.

        Args:
            image_path: Path to image file
            sheet_name: Name of the sheet
            cell: Cell location for image (e.g., "A1")
            width: Optional width in pixels
            height: Optional height in pixels
        """
        if not self.workbook:
            self._ensure_workbook()

        image_path = Path(image_path)
        if not image_path.exists():
            execution_logger.logger.error(f"Image not found: {image_path}")
            return

        # Get or create sheet
        if sheet_name not in self.workbook.sheetnames:
            ws = self.workbook.create_sheet(sheet_name)
        else:
            ws = self.workbook[sheet_name]

        # Process image
        try:
            # Open and potentially resize image
            pil_image = Image.open(image_path)

            if width or height:
                # Calculate aspect ratio
                aspect_ratio = pil_image.width / pil_image.height

                if width and not height:
                    height = int(width / aspect_ratio)
                elif height and not width:
                    width = int(height * aspect_ratio)

                # Resize image
                pil_image = pil_image.resize((width or pil_image.width, height or pil_image.height))

                # Save resized image to temporary file
                import tempfile
                temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                pil_image.save(temp_file.name, format="PNG")
                temp_file.close()

                # Create openpyxl image from temp file
                img = OpenpyxlImage(temp_file.name)
            else:
                # Use original image
                img = OpenpyxlImage(str(image_path))

            # Add image to worksheet
            ws.add_image(img, cell)

            execution_logger.log_action(
                "embed_image",
                sheet_name,
                image_path=str(image_path),
                cell=cell,
            )

        except Exception as e:
            execution_logger.logger.error(f"Failed to embed image: {e}")

    def create_sheet(self, sheet_name: str) -> Worksheet:
        """
        Create a new sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Created worksheet
        """
        if not self.workbook:
            self._ensure_workbook()

        # Remove if exists
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])

        ws = self.workbook.create_sheet(sheet_name)
        self.save()

        execution_logger.log_action("create_sheet", sheet_name)
        return ws

    def write_dict_list(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str,
        header_style: bool = True,
    ) -> None:
        """
        Write list of dictionaries to sheet.

        Args:
            data: List of dictionaries
            sheet_name: Name of the sheet
            header_style: Whether to apply header styling
        """
        if not data:
            # Create empty sheet
            self.create_sheet(sheet_name)
            return

        # Convert to DataFrame
        df = pd.DataFrame(data)
        self.write_df(df, sheet_name, header_style=header_style)

    def _autosize_columns(self, ws: Worksheet) -> None:
        """
        Autosize columns based on content.

        Args:
            ws: Worksheet to autosize
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def save(self) -> None:
        """Save workbook to file."""
        if self.workbook:
            self.workbook.save(self.filepath)

    def close(self) -> None:
        """Close workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        if not self.workbook:
            self._ensure_workbook()
        return list(self.workbook.sheetnames) if self.workbook else []

    def sheet_exists(self, sheet_name: str) -> bool:
        """Check if sheet exists."""
        return sheet_name in self.get_sheet_names()

    def read_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """
        Read sheet as DataFrame.

        Args:
            sheet_name: Name of the sheet

        Returns:
            DataFrame or None if sheet doesn't exist
        """
        if not self.sheet_exists(sheet_name):
            return None

        return pd.read_excel(self.filepath, sheet_name=sheet_name)


# Convenience functions
def ensure_workbook(filepath: Optional[Union[str, Path]] = None) -> ExcelWriter:
    """Ensure workbook exists and return writer."""
    return ExcelWriter(filepath)


def write_df(df: pd.DataFrame, sheet_name: str, filepath: Optional[Union[str, Path]] = None) -> None:
    """Write DataFrame to sheet."""
    writer = ExcelWriter(filepath)
    writer.write_df(df, sheet_name)


def append_df(df: pd.DataFrame, sheet_name: str, filepath: Optional[Union[str, Path]] = None) -> None:
    """Append DataFrame to sheet."""
    writer = ExcelWriter(filepath)
    writer.append_df(df, sheet_name)


def embed_image(
    image_path: Union[str, Path],
    sheet_name: str,
    cell: str = "A1",
    filepath: Optional[Union[str, Path]] = None,
) -> None:
    """Embed image in worksheet."""
    writer = ExcelWriter(filepath)
    writer.embed_image(image_path, sheet_name, cell)
